import logging
from datetime import datetime, timezone

from pymongo import MongoClient, DESCENDING
from pymongo.collection import Collection
from pymongo.errors import ConnectionFailure, PyMongoError
from config import config_manager

logger = logging.getLogger(__name__)


class DatabasePool:
    """Persistent MongoDB connection managed by PyMongo's MongoClient.

    Handles connection failures gracefully at startup — a warning is logged
    instead of crashing the server if MongoDB is temporarily unavailable.
    MongoClient maintains an internal connection pool and is thread-safe.
    """

    def __init__(self, uri: str, db_name: str):
        self.uri = uri
        self.db_name = db_name
        self._client: MongoClient | None = None
        self._db = None
        self._initialize_client()

    def _initialize_client(self) -> None:
        try:
            self._client = MongoClient(
                self.uri,
                serverSelectionTimeoutMS=5000,
                connectTimeoutMS=5000,
                socketTimeoutMS=30000,
                maxPoolSize=50,
                minPoolSize=1,
            )
            self._client.admin.command("ping")
            self._db = self._client[self.db_name]
            logger.info("MongoDB connected: %s/%s", self.uri, self.db_name)
        except (ConnectionFailure, PyMongoError) as e:
            logger.warning("MongoDB unavailable at startup: %s", e)
            self._client = None
            self._db = None

    def get_collection(self, name: str) -> Collection:
        if self._db is None:
            raise RuntimeError("MongoDB is not connected")
        return self._db[name]

    def ping(self) -> bool:
        try:
            if self._client is None:
                return False
            self._client.admin.command("ping")
            return True
        except (ConnectionFailure, PyMongoError):
            return False

    def close(self) -> None:
        if self._client is not None:
            self._client.close()
            self._client = None
            self._db = None


_pool = DatabasePool(config_manager.db.mongo_uri, config_manager.db.db_name)


def _col_events() -> Collection:
    return _pool.get_collection(config_manager.db.events_collection)


def _col_children() -> Collection:
    return _pool.get_collection("children")


def _col_words() -> Collection:
    return _pool.get_collection(config_manager.db.words_collection)


def initialize_indexes() -> None:
    """Create MongoDB indexes. Call once from the server lifespan."""
    try:
        _col_events().create_index([("childId", 1), ("timestamp", DESCENDING)])
        _col_children().create_index("childId", unique=True)
        _col_words().create_index("word", unique=True)
        logger.info("MongoDB indexes ensured")
    except (RuntimeError, PyMongoError) as e:
        logger.warning("Could not create indexes: %s", e)


def ping() -> bool:
    """Returns True if the MongoDB connection is healthy."""
    return _pool.ping()


def insert_event(doc: dict) -> str:
    """Insert a hop event. Returns the inserted document's string id."""
    result = _col_events().insert_one(doc)
    return str(result.inserted_id)


def get_events(child_id: str) -> list[dict]:
    """Return all events for a child, newest first, with _id serialised to str."""
    cursor = _col_events().find({"childId": child_id}, {"_id": 0}).sort("timestamp", DESCENDING)
    return list(cursor)


def clear_events(child_id: str) -> int:
    """Delete all hop events for a child. Returns the number of deleted documents."""
    result = _col_events().delete_many({"childId": child_id})
    return result.deleted_count


def register_child(child_id: str, child_name: str) -> None:
    """Insert a child record only if it doesn't exist yet. Never overwrites the stored name."""
    _col_children().update_one(
        {"childId": child_id},
        {"$setOnInsert": {
            "childName": child_name,
            "registeredAt": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )


def rename_child(child_id: str, child_name: str) -> bool:
    """Update the stored name for an existing child. Returns True if a doc was modified."""
    result = _col_children().update_one(
        {"childId": child_id},
        {"$set": {"childName": child_name}},
    )
    return result.modified_count > 0


def get_children() -> list[dict]:
    """Return all registered children, falling back to event-derived ones."""
    registered = {
        doc["childId"]: doc.get("childName", doc["childId"])
        for doc in _col_children().find({}, {"_id": 0, "childId": 1, "childName": 1})
    }
    # Also include any child that has events but was never explicitly registered.
    for cid in _col_events().distinct("childId"):
        registered.setdefault(cid, cid)
    return [{"childId": cid, "childName": name} for cid, name in sorted(registered.items())]


# ---------------------------------------------------------------------------
# Blocked words
# ---------------------------------------------------------------------------

def get_blocked_words() -> set[str]:
    """Return the full blocked-words set from MongoDB."""
    return {doc["word"] for doc in _col_words().find({}, {"_id": 0, "word": 1})}


def add_word(word: str) -> bool:
    """Upsert a blocked word. Returns True if the word was newly inserted."""
    from pymongo.errors import DuplicateKeyError
    try:
        _col_words().insert_one({"word": word, "addedAt": datetime.now(timezone.utc).isoformat()})
        return True
    except DuplicateKeyError:
        return False


def remove_word(word: str) -> bool:
    """Delete a blocked word. Returns True if it existed."""
    result = _col_words().delete_one({"word": word})
    return result.deleted_count > 0


def seed_words_from_excel(path: str) -> int:
    """Bulk-upsert words from an Excel file (column: 'word') into MongoDB.

    Safe to call multiple times — uses upsert so existing words are ignored.
    Returns the number of words processed (not just newly inserted).
    """
    import os
    import openpyxl
    from pymongo import UpdateOne

    if not os.path.exists(path):
        logger.warning("seed_words_from_excel: file not found at %r", path)
        return 0

    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        word_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() == "word":
                word_col_idx = cell.column - 1
                break

        if word_col_idx is None:
            logger.warning("seed_words_from_excel: no 'word' column found in %r", path)
            wb.close()
            return 0

        ops = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[word_col_idx]:
                w = str(row[word_col_idx]).strip().lower()
                if w:
                    ops.append(UpdateOne(
                        {"word": w},
                        {"$setOnInsert": {"word": w, "addedAt": datetime.now(timezone.utc).isoformat()}},
                        upsert=True,
                    ))
        wb.close()

        if ops:
            _col_words().bulk_write(ops, ordered=False)

        logger.info("seed_words_from_excel: processed %d words from %r", len(ops), path)
        return len(ops)

    except Exception as exc:
        logger.warning("seed_words_from_excel failed: %s", exc)
        return 0

