import logging
import os
from datetime import datetime, timezone

import openpyxl
from pymongo.errors import DuplicateKeyError
from pymongo import UpdateOne

from core.database import pool
from config import config_manager

logger = logging.getLogger(__name__)


def _col_words():
    return pool.get_collection(config_manager.db.words_collection)


def get_blocked_words() -> set[str]:
    return {doc["word"] for doc in _col_words().find({}, {"_id": 0, "word": 1})}


def has_words() -> bool:
    return _col_words().count_documents({}, limit=1) > 0


def add_word(word: str) -> bool:
    try:
        _col_words().insert_one({"word": word, "addedAt": datetime.now(timezone.utc).isoformat()})
        return True
    except DuplicateKeyError:
        return False


def remove_word(word: str) -> bool:
    result = _col_words().delete_one({"word": word})
    return result.deleted_count > 0


def seed_words_from_excel(path: str) -> int:
    if not os.path.exists(path):
        logger.warning("seed_words_from_excel: file not found at %r", path)
        return 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        word_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() == "word":
                word_col_idx = cell.column - 1
                break
        if word_col_idx is None:
            logger.warning("seed_words_from_excel: no 'word' column found in %r", path)
            wb.close()
            return 0
        ops = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[word_col_idx]:
                w = str(row[word_col_idx]).strip().lower()
                if w:
                    ops.append(UpdateOne(
                        {"word": w},
                        {"$setOnInsert": {"word": w, "addedAt": datetime.now(timezone.utc).isoformat()}},
                        upsert=True,
                    ))
        wb.close()
        if ops:
            _col_words().bulk_write(ops, ordered=False)
        logger.info("seed_words_from_excel: processed %d words from %r", len(ops), path)
        return len(ops)
    except Exception as exc:
        logger.warning("seed_words_from_excel failed: %s", exc)
        return 0


def initialize_indexes() -> None:
    _col_words().create_index("word", unique=True)
