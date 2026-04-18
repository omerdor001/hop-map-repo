"""
Unit tests for check_blocked_words() via words.service.

These tests inject a WordsFilter directly — no server process, no HTTP calls,
no database.
"""

import pytest
import openpyxl

import sys
from pathlib import Path

_SERVER_DIR = Path(__file__).resolve().parent.parent.parent
if str(_SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(_SERVER_DIR))

import words.service as words_service
import words.repository as words_repo
from words.service import check_blocked_words
from words_filter import WordsFilter


def _set_words(monkeypatch, words: set[str]) -> None:
    f = WordsFilter()
    f.build(words)
    monkeypatch.setattr(words_service, "_filter", f)


class TestCheckBlockedPhrases:

    def test_detects_exact_phrase(self, monkeypatch):
        _set_words(monkeypatch, {"don't tell your parents"})
        found, matched = check_blocked_words("hey man don't tell your parents about this")
        assert found is True
        assert matched == "don't tell your parents"

    def test_phrase_match_is_case_insensitive(self, monkeypatch):
        _set_words(monkeypatch, {"join my server"})
        found, matched = check_blocked_words("COME ON JOIN MY SERVER RIGHT NOW")
        assert found is True
        assert matched == "join my server"

    def test_phrase_not_matched_when_words_are_split(self, monkeypatch):
        _set_words(monkeypatch, {"join my server"})
        found, _ = check_blocked_words("join the my server today")
        assert found is False

    def test_phrase_with_inserted_punctuation_not_matched(self, monkeypatch):
        _set_words(monkeypatch, {"send nudes"})
        found, _ = check_blocked_words("send, nudes")
        assert found is False

    def test_longest_phrase_matched_first(self, monkeypatch):
        _set_words(monkeypatch, {"tell your parents", "don't tell your parents"})
        found, matched = check_blocked_words("please don't tell your parents about this")
        assert found is True
        assert matched == "don't tell your parents"

    def test_empty_phrases_returns_false(self, monkeypatch):
        _set_words(monkeypatch, set())
        found, matched = check_blocked_words("don't tell your parents")
        assert found is False
        assert matched == ""

    def test_phrase_and_word_both_present_phrase_matched_first(self, monkeypatch):
        _set_words(monkeypatch, {"send nudes", "hack"})
        found, matched = check_blocked_words("hack me and send nudes")
        assert found is True
        assert matched == "send nudes"


class TestCheckBlockedWords:

    def test_returns_false_when_word_set_is_empty(self, monkeypatch):
        _set_words(monkeypatch, set())
        found, matched = check_blocked_words("hack the planet")
        assert found is False
        assert matched == ""

    def test_detects_exact_blocked_word(self, monkeypatch):
        _set_words(monkeypatch, {"hack"})
        found, matched = check_blocked_words("click here to hack your account")
        assert found is True
        assert matched == "hack"

    def test_case_insensitive_match(self, monkeypatch):
        _set_words(monkeypatch, {"hack"})
        found, matched = check_blocked_words("HACK the system")
        assert found is True
        assert matched == "hack"

    def test_clean_phrase_not_flagged(self, monkeypatch):
        _set_words(monkeypatch, {"hack", "kill", "nude"})
        found, _ = check_blocked_words("hey bro come check this cool server")
        assert found is False

    def test_blocked_word_inside_longer_word_not_matched(self, monkeypatch):
        _set_words(monkeypatch, {"hack"})
        found, _ = check_blocked_words("hackle is a word about feathers")
        assert found is False

    def test_first_match_in_text_order_is_returned(self, monkeypatch):
        _set_words(monkeypatch, {"kill", "nude", "hack"})
        found, matched = check_blocked_words("hack and nude content")
        assert found is True
        assert matched == "hack"

    def test_empty_string_returns_false(self, monkeypatch):
        _set_words(monkeypatch, {"hack"})
        found, matched = check_blocked_words("")
        assert found is False
        assert matched == ""

    def test_whitespace_only_returns_false(self, monkeypatch):
        _set_words(monkeypatch, {"hack"})
        found, _ = check_blocked_words("   \t\n  ")
        assert found is False

    def test_word_with_surrounding_punctuation_still_detected(self, monkeypatch):
        _set_words(monkeypatch, {"hack"})
        found, matched = check_blocked_words("don't hack, please!")
        assert found is True
        assert matched == "hack"

    def test_returns_false_not_none(self, monkeypatch):
        _set_words(monkeypatch, set())
        found, matched = check_blocked_words("anything")
        assert found is False
        assert matched == ""


class TestCheckBlockedWordsWithRealDB:
    """Spot-check known entries against the real hopmap_words_db.xlsx.
    Skipped automatically if the file is not present.
    """

    @pytest.fixture(autouse=True)
    def _load_real_words(self, monkeypatch, words_db_path):
        wb = openpyxl.load_workbook(words_db_path, read_only=True)
        ws = wb.active
        word_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() == "word":
                word_col_idx = cell.column - 1
                break
        words: set[str] = set()
        if word_col_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[word_col_idx]:
                    w = str(row[word_col_idx]).strip().lower()
                    if w:
                        words.add(w)
        wb.close()

        monkeypatch.setattr(words_repo, "get_blocked_words", lambda: words)
        _saved_filter = words_service._filter
        words_service.load_blocked_words()
        yield
        words_service._filter = _saved_filter

    def test_hack_is_blocked(self):
        found, matched = check_blocked_words("click here to hack your friends account")
        assert found is True
        assert matched == "hack"

    def test_grooming_phrase_is_blocked(self):
        found, matched = check_blocked_words("this is our secret don't tell your parents ok?")
        assert found is True
        assert "parents" in matched

    def test_clean_phrase_passes(self):
        found, _ = check_blocked_words("hey bro come check this cool server")
        assert found is False

    def test_18plus_blocked(self):
        found, _ = check_blocked_words("check out this 18+ content")
        assert found is True
