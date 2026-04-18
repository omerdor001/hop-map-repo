"""Unit tests for the repository layer using mongomock as an in-memory MongoDB substitute.

All tests patch pymongo.MongoClient with mongomock before the pool initialises so
that no real MongoDB instance is required.
"""
from __future__ import annotations

import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import patch

import mongomock
import openpyxl
import pytest
from pymongo.errors import DuplicateKeyError

_SERVER_DIR = Path(__file__).resolve().parent.parent.parent
if str(_SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(_SERVER_DIR))


# ---------------------------------------------------------------------------
# Module-level fixture — wire mongomock into the shared pool
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def db_mod():
    """Redirect the shared DatabasePool to a mongomock client, then expose all
    repository functions through a single facade so test code reads identically
    to the old monolithic db.py tests."""
    mock_client = mongomock.MongoClient()

    import core.database as _core_db
    _core_db._pool._client = mock_client
    _core_db._pool._db = mock_client[_core_db._pool.db_name]

    import auth.repository as auth_repo
    import children.repository as child_repo
    import events.repository as event_repo
    import notifications.repository as notif_repo
    import words.repository as words_repo

    try:
        auth_repo.initialize_indexes()
        child_repo.initialize_indexes()
        event_repo.initialize_indexes()
        notif_repo.initialize_indexes()
        words_repo.initialize_indexes()
    except Exception:
        pass

    class DbFacade:
        # events
        insert_event          = staticmethod(event_repo.insert_event)
        get_events            = staticmethod(event_repo.get_events)
        clear_events          = staticmethod(event_repo.clear_events)
        # children
        register_child        = staticmethod(child_repo.register_child)
        rename_child          = staticmethod(child_repo.rename_child)
        get_children          = staticmethod(child_repo.get_children)
        get_child_by_id       = staticmethod(child_repo.get_child_by_id)
        get_child_by_agent_token = staticmethod(child_repo.get_child_by_agent_token)
        # words
        get_blocked_words     = staticmethod(words_repo.get_blocked_words)
        add_word              = staticmethod(words_repo.add_word)
        remove_word           = staticmethod(words_repo.remove_word)
        seed_words_from_excel = staticmethod(words_repo.seed_words_from_excel)
        # users
        create_user           = staticmethod(auth_repo.create_user)
        get_user_by_email     = staticmethod(auth_repo.get_user_by_email)
        get_user_by_id        = staticmethod(auth_repo.get_user_by_id)
        # sessions
        create_session        = staticmethod(auth_repo.create_session)
        get_session_by_hash   = staticmethod(auth_repo.get_session_by_hash)
        revoke_session        = staticmethod(auth_repo.revoke_session)
        # notifications
        insert_notification   = staticmethod(notif_repo.insert_notification)
        get_notifications     = staticmethod(notif_repo.get_notifications)
        mark_notification_read = staticmethod(notif_repo.mark_notification_read)

        @staticmethod
        def _col_events():
            from core.database import _pool
            from config import config_manager
            return _pool.get_collection(config_manager.db.events_collection)

        @staticmethod
        def _col_children():
            from core.database import _pool
            return _pool.get_collection("children")

        @staticmethod
        def _col_words():
            from core.database import _pool
            from config import config_manager
            return _pool.get_collection(config_manager.db.words_collection)

        @staticmethod
        def _col_users():
            from core.database import _pool
            return _pool.get_collection("users")

        @staticmethod
        def _col_sessions():
            from core.database import _pool
            return _pool.get_collection("sessions")

        @staticmethod
        def _col_notifications():
            from core.database import _pool
            return _pool.get_collection("notifications")

        @staticmethod
        def ping():
            from core.database import _pool
            return _pool.ping()

        @property
        def _pool(self):
            from core.database import _pool
            return _pool

    return DbFacade()


@pytest.fixture(autouse=True)
def clean_collections(db_mod):
    for col in (
        db_mod._col_events(),
        db_mod._col_children(),
        db_mod._col_words(),
        db_mod._col_users(),
        db_mod._col_sessions(),
        db_mod._col_notifications(),
    ):
        col.delete_many({})
    yield


# ---------------------------------------------------------------------------
# Test helpers
# ---------------------------------------------------------------------------

_PARENT_A = "parent-aaa"
_PARENT_B = "parent-bbb"


def _register(db_mod, child_id: str, name: str, parent_id: str = _PARENT_A) -> None:
    db_mod.register_child(child_id, name, parent_id, f"hash-{child_id}", child_id[:8])


def _make_user(db_mod, email: str = "test@example.com", display_name: str = "Test User") -> str:
    return db_mod.create_user(email, "fake-bcrypt-hash", display_name)


def _future(days: int = 30) -> datetime:
    return datetime.now(timezone.utc) + timedelta(days=days)


def _past(seconds: int = 1) -> datetime:
    return datetime.now(timezone.utc) - timedelta(seconds=seconds)


# ---------------------------------------------------------------------------
# Events
# ---------------------------------------------------------------------------

class TestInsertAndGetEvents:

    def test_insert_event_returns_string_id(self, db_mod):
        doc = {"childId": "c1", "from": "roblox.exe", "to": "discord.exe",
               "timestamp": datetime.now(timezone.utc).isoformat()}
        result = db_mod.insert_event(doc)
        assert isinstance(result, str) and len(result) > 0

    def test_get_events_returns_empty_list_for_unknown_child(self, db_mod):
        assert db_mod.get_events("nonexistent") == []

    def test_get_events_returns_all_inserted_events(self, db_mod):
        db_mod.insert_event({"childId": "c2", "timestamp": "2026-01-01T00:00:00Z"})
        db_mod.insert_event({"childId": "c2", "timestamp": "2026-01-02T00:00:00Z"})
        assert len(db_mod.get_events("c2")) == 2

    def test_get_events_does_not_leak_other_children(self, db_mod):
        db_mod.insert_event({"childId": "c3", "timestamp": "2026-01-01T00:00:00Z"})
        db_mod.insert_event({"childId": "other", "timestamp": "2026-01-01T00:00:00Z"})
        assert len(db_mod.get_events("c3")) == 1


class TestClearEvents:

    def test_returns_deleted_count(self, db_mod):
        db_mod.insert_event({"childId": "c4", "timestamp": "2026-01-01T00:00:00Z"})
        db_mod.insert_event({"childId": "c4", "timestamp": "2026-01-02T00:00:00Z"})
        assert db_mod.clear_events("c4") == 2

    def test_empties_collection_for_child(self, db_mod):
        db_mod.insert_event({"childId": "c5", "timestamp": "2026-01-01T00:00:00Z"})
        db_mod.clear_events("c5")
        assert db_mod.get_events("c5") == []

    def test_does_not_affect_other_children(self, db_mod):
        db_mod.insert_event({"childId": "c6", "timestamp": "2026-01-01T00:00:00Z"})
        db_mod.insert_event({"childId": "c7", "timestamp": "2026-01-01T00:00:00Z"})
        db_mod.clear_events("c6")
        assert len(db_mod.get_events("c7")) == 1

    def test_on_empty_child_returns_zero(self, db_mod):
        assert db_mod.clear_events("nobody") == 0


# ---------------------------------------------------------------------------
# Children
# ---------------------------------------------------------------------------

class TestRegisterChild:

    def test_child_appears_in_get_children_for_its_parent(self, db_mod):
        _register(db_mod, "kid-1", "Alice", _PARENT_A)
        children = {c["childId"]: c["childName"] for c in db_mod.get_children(_PARENT_A)}
        assert children.get("kid-1") == "Alice"

    def test_child_is_not_visible_to_another_parent(self, db_mod):
        _register(db_mod, "kid-1", "Alice", _PARENT_A)
        assert db_mod.get_children(_PARENT_B) == []

    def test_register_does_not_overwrite_existing_name(self, db_mod):
        _register(db_mod, "kid-2", "Bob", _PARENT_A)
        _register(db_mod, "kid-2", "NewBob", _PARENT_A)
        children = {c["childId"]: c["childName"] for c in db_mod.get_children(_PARENT_A)}
        assert children["kid-2"] == "Bob"

    def test_agent_token_hash_is_stored(self, db_mod):
        db_mod.register_child("kid-3", "Carol", _PARENT_A, "tok-hash-abc", "tok-hash")
        doc = db_mod._col_children().find_one({"childId": "kid-3"})
        assert doc["agentTokenHash"] == "tok-hash-abc"

    def test_parent_id_is_stored(self, db_mod):
        _register(db_mod, "kid-4", "Dave", _PARENT_A)
        doc = db_mod._col_children().find_one({"childId": "kid-4"})
        assert doc["parentId"] == _PARENT_A


class TestGetChildById:

    def test_returns_child_for_correct_parent(self, db_mod):
        _register(db_mod, "kid-5", "Eve", _PARENT_A)
        doc = db_mod.get_child_by_id("kid-5", _PARENT_A)
        assert doc is not None and doc["childId"] == "kid-5"

    def test_returns_none_for_wrong_parent(self, db_mod):
        _register(db_mod, "kid-6", "Frank", _PARENT_A)
        assert db_mod.get_child_by_id("kid-6", _PARENT_B) is None

    def test_returns_none_for_unknown_child(self, db_mod):
        assert db_mod.get_child_by_id("nobody", _PARENT_A) is None


class TestGetChildByAgentToken:

    def test_returns_child_for_matching_hash(self, db_mod):
        db_mod.register_child("kid-7", "Grace", _PARENT_A, "exact-hash-xyz", "exact-ha")
        doc = db_mod.get_child_by_agent_token("exact-hash-xyz")
        assert doc is not None and doc["childId"] == "kid-7"

    def test_returns_none_for_unknown_hash(self, db_mod):
        assert db_mod.get_child_by_agent_token("nonexistent-hash") is None


class TestRenameChild:

    def test_returns_true_on_success(self, db_mod):
        _register(db_mod, "kid-8", "Hank", _PARENT_A)
        assert db_mod.rename_child("kid-8", "Hank 2") is True

    def test_name_is_updated(self, db_mod):
        _register(db_mod, "kid-9", "Iris", _PARENT_A)
        db_mod.rename_child("kid-9", "Iris 2")
        children = {c["childId"]: c["childName"] for c in db_mod.get_children(_PARENT_A)}
        assert children["kid-9"] == "Iris 2"

    def test_returns_false_for_unknown_id(self, db_mod):
        assert db_mod.rename_child("nobody", "Ghost") is False


# ---------------------------------------------------------------------------
# Blocked words
# ---------------------------------------------------------------------------

class TestBlockedWords:

    def test_initially_empty(self, db_mod):
        assert db_mod.get_blocked_words() == set()

    def test_add_new_word_returns_true(self, db_mod):
        assert db_mod.add_word("hack") is True

    def test_added_word_appears_in_get_blocked_words(self, db_mod):
        db_mod.add_word("discord")
        assert "discord" in db_mod.get_blocked_words()

    def test_add_duplicate_word_returns_false(self, db_mod):
        db_mod.add_word("telegram")
        assert db_mod.add_word("telegram") is False

    def test_remove_existing_word_returns_true(self, db_mod):
        db_mod.add_word("snap")
        assert db_mod.remove_word("snap") is True

    def test_removed_word_absent_from_get_blocked_words(self, db_mod):
        db_mod.add_word("signal")
        db_mod.remove_word("signal")
        assert "signal" not in db_mod.get_blocked_words()

    def test_remove_nonexistent_word_returns_false(self, db_mod):
        assert db_mod.remove_word("ghost") is False


# ---------------------------------------------------------------------------
# Users
# ---------------------------------------------------------------------------

class TestCreateUser:

    def test_returns_string_id(self, db_mod):
        user_id = _make_user(db_mod, "a@example.com")
        assert isinstance(user_id, str) and len(user_id) > 0

    def test_email_is_normalised_to_lowercase(self, db_mod):
        _make_user(db_mod, "Upper@EXAMPLE.COM")
        doc = db_mod._col_users().find_one({"email": "upper@example.com"})
        assert doc is not None

    def test_duplicate_email_raises_duplicate_key_error(self, db_mod):
        _make_user(db_mod, "dup@example.com")
        with pytest.raises(DuplicateKeyError):
            _make_user(db_mod, "dup@example.com")

    def test_deleted_at_is_none_on_creation(self, db_mod):
        _make_user(db_mod, "new@example.com")
        doc = db_mod._col_users().find_one({"email": "new@example.com"})
        assert doc["deletedAt"] is None


class TestGetUserByEmail:

    def test_returns_user_for_known_email(self, db_mod):
        _make_user(db_mod, "known@example.com")
        user = db_mod.get_user_by_email("known@example.com")
        assert user is not None and user["email"] == "known@example.com"

    def test_returns_none_for_unknown_email(self, db_mod):
        assert db_mod.get_user_by_email("nobody@example.com") is None

    def test_returned_doc_has_id_not_underscore_id(self, db_mod):
        _make_user(db_mod, "id-check@example.com")
        user = db_mod.get_user_by_email("id-check@example.com")
        assert "id" in user and "_id" not in user

    def test_lookup_is_case_insensitive(self, db_mod):
        _make_user(db_mod, "mixed@example.com")
        assert db_mod.get_user_by_email("MIXED@EXAMPLE.COM") is not None

    def test_does_not_return_soft_deleted_user(self, db_mod):
        _make_user(db_mod, "deleted@example.com")
        db_mod._col_users().update_one(
            {"email": "deleted@example.com"},
            {"$set": {"deletedAt": datetime.now(timezone.utc).isoformat()}},
        )
        assert db_mod.get_user_by_email("deleted@example.com") is None


class TestGetUserById:

    def test_returns_user_for_valid_id(self, db_mod):
        user_id = _make_user(db_mod, "by-id@example.com")
        user = db_mod.get_user_by_id(user_id)
        assert user is not None and user["id"] == user_id

    def test_returns_none_for_invalid_object_id_string(self, db_mod):
        assert db_mod.get_user_by_id("not-an-objectid") is None

    def test_returns_none_for_unknown_valid_id(self, db_mod):
        assert db_mod.get_user_by_id("000000000000000000000000") is None

    def test_does_not_return_soft_deleted_user(self, db_mod):
        user_id = _make_user(db_mod, "del-by-id@example.com")
        db_mod._col_users().update_one(
            {"email": "del-by-id@example.com"},
            {"$set": {"deletedAt": datetime.now(timezone.utc).isoformat()}},
        )
        assert db_mod.get_user_by_id(user_id) is None


# ---------------------------------------------------------------------------
# Sessions
# ---------------------------------------------------------------------------

class TestSessions:

    def test_create_session_returns_string_id(self, db_mod):
        assert isinstance(db_mod.create_session("uid-1", "hash-abc", _future()), str)

    def test_get_session_by_hash_returns_session_with_correct_user(self, db_mod):
        db_mod.create_session("uid-1", "hash-xyz", _future())
        session = db_mod.get_session_by_hash("hash-xyz")
        assert session is not None and session["userId"] == "uid-1"

    def test_get_session_returns_none_for_unknown_hash(self, db_mod):
        assert db_mod.get_session_by_hash("unknown-hash") is None

    def test_get_session_returns_none_for_expired_session(self, db_mod):
        db_mod.create_session("uid-1", "hash-exp", _past())
        assert db_mod.get_session_by_hash("hash-exp") is None

    def test_revoke_session_makes_it_unfindable(self, db_mod):
        db_mod.create_session("uid-1", "hash-rev", _future())
        db_mod.revoke_session("hash-rev")
        assert db_mod.get_session_by_hash("hash-rev") is None

    def test_revoking_one_session_does_not_affect_others(self, db_mod):
        db_mod.create_session("uid-1", "hash-keep", _future())
        db_mod.create_session("uid-1", "hash-kill", _future())
        db_mod.revoke_session("hash-kill")
        assert db_mod.get_session_by_hash("hash-keep") is not None

    def test_revoke_nonexistent_session_does_not_raise(self, db_mod):
        db_mod.revoke_session("nonexistent-hash")


# ---------------------------------------------------------------------------
# Notifications
# ---------------------------------------------------------------------------

class TestInsertNotification:

    def test_returns_string_id(self, db_mod):
        nid = db_mod.insert_notification(_PARENT_A, "child-1", "evt-1", "hop_detected", "Alice hopped")
        assert isinstance(nid, str) and len(nid) > 0

    def test_notification_is_unread_by_default(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "child-1", "evt-1", "hop_detected", "Alice hopped")
        doc = db_mod._col_notifications().find_one({"parentId": _PARENT_A})
        assert doc["read"] is False

    def test_notification_stores_correct_fields(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "child-1", "evt-123", "hop_detected", "Hello")
        doc = db_mod._col_notifications().find_one({"parentId": _PARENT_A})
        assert doc["childId"] == "child-1"
        assert doc["eventId"] == "evt-123"
        assert doc["type"] == "hop_detected"
        assert doc["message"] == "Hello"


class TestGetNotifications:

    def test_returns_all_notifications_for_parent(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "msg1")
        db_mod.insert_notification(_PARENT_A, "c1", "e2", "hop_detected", "msg2")
        assert len(db_mod.get_notifications(_PARENT_A)) == 2

    def test_does_not_return_other_parents_notifications(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "for A")
        db_mod.insert_notification(_PARENT_B, "c2", "e2", "hop_detected", "for B")
        results = db_mod.get_notifications(_PARENT_A)
        assert len(results) == 1 and results[0]["message"] == "for A"

    def test_unread_only_excludes_read_notifications(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "unread")
        nid = db_mod.insert_notification(_PARENT_A, "c1", "e2", "hop_detected", "read")
        db_mod.mark_notification_read(nid, _PARENT_A)
        results = db_mod.get_notifications(_PARENT_A, unread_only=True)
        assert len(results) == 1 and results[0]["message"] == "unread"

    def test_returns_empty_list_when_no_notifications(self, db_mod):
        assert db_mod.get_notifications(_PARENT_A) == []

    def test_returned_docs_have_id_not_underscore_id(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "msg")
        doc = db_mod.get_notifications(_PARENT_A)[0]
        assert "id" in doc and "_id" not in doc

    def test_created_at_is_serialised_to_iso_string(self, db_mod):
        db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "msg")
        doc = db_mod.get_notifications(_PARENT_A)[0]
        assert isinstance(doc["createdAt"], str)


class TestMarkNotificationRead:

    def test_notification_is_read_after_marking(self, db_mod):
        nid = db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "msg")
        db_mod.mark_notification_read(nid, _PARENT_A)
        assert db_mod.get_notifications(_PARENT_A, unread_only=True) == []

    def test_returns_true_on_success(self, db_mod):
        nid = db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "msg")
        assert db_mod.mark_notification_read(nid, _PARENT_A) is True

    def test_returns_false_for_invalid_object_id(self, db_mod):
        assert db_mod.mark_notification_read("not-an-objectid", _PARENT_A) is False

    def test_returns_false_for_wrong_parent(self, db_mod):
        nid = db_mod.insert_notification(_PARENT_A, "c1", "e1", "hop_detected", "msg")
        assert db_mod.mark_notification_read(nid, _PARENT_B) is False


# ---------------------------------------------------------------------------
# DatabasePool
# ---------------------------------------------------------------------------

class TestDatabasePool:

    def test_ping_returns_true_when_connected(self, db_mod):
        assert db_mod.ping() is True

    def test_client_is_not_none_after_successful_init(self, db_mod):
        assert db_mod._pool._client is not None


# ---------------------------------------------------------------------------
# seed_words_from_excel
# ---------------------------------------------------------------------------

class TestSeedWordsFromExcel:

    @pytest.fixture(autouse=True)
    def _patch_update_one_sort(self, monkeypatch):
        # mongomock 4.3 doesn't accept the 'sort' kwarg pymongo 4.4+ passes through bulk_write
        from mongomock.collection import BulkOperationBuilder
        _orig = BulkOperationBuilder.add_update
        def _patched(self, selector, doc, multi=False, upsert=False,
                     collation=None, array_filters=None, hint=None, **kwargs):
            kwargs.pop("sort", None)
            return _orig(self, selector, doc, multi=multi, upsert=upsert,
                         collation=collation, array_filters=array_filters, hint=hint)
        monkeypatch.setattr(BulkOperationBuilder, "add_update", _patched)

    def test_missing_file_returns_zero(self, db_mod):
        assert db_mod.seed_words_from_excel("/nonexistent/path/file.xlsx") == 0

    def test_valid_excel_inserts_words(self, db_mod, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["word"])
        for w in ["badword1", "badword2", "badword3"]:
            ws.append([w])
        xlsx_path = tmp_path / "words.xlsx"
        wb.save(str(xlsx_path))
        count = db_mod.seed_words_from_excel(str(xlsx_path))
        assert count == 3
        assert {"badword1", "badword2", "badword3"}.issubset(db_mod.get_blocked_words())

    def test_excel_without_word_column_returns_zero(self, db_mod, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["token"])
        ws.append(["something"])
        xlsx_path = tmp_path / "no_word_col.xlsx"
        wb.save(str(xlsx_path))
        assert db_mod.seed_words_from_excel(str(xlsx_path)) == 0

    def test_seed_is_idempotent(self, db_mod, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["word"])
        ws.append(["unique1"])
        xlsx_path = tmp_path / "idem.xlsx"
        wb.save(str(xlsx_path))
        db_mod.seed_words_from_excel(str(xlsx_path))
        db_mod.seed_words_from_excel(str(xlsx_path))
        assert list(db_mod.get_blocked_words()).count("unique1") == 1
